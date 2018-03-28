import time

from dataBaseConnect import ConnecyMyDataBase
from openpyxl import load_workbook

inicio = time.time()
print("Lendo o arquivo")
wb = load_workbook('/home/ldrf/Projetos/TCC/Planilha/Q eu acho/Tabela IES 2017.xlsx', read_only=True)
data = ''
print("Lendo as folhas")
for sheet_name in wb.sheetnames:
    print(sheet_name)

sheet = wb[wb.sheetnames[0]]
data = ""
first_line = ""
database = ConnecyMyDataBase()
print('Conectando com o banco de dados')
database.set_connection('uece', '1', '127.0.0.1', database='EGRESSO')
print('Conexao com sucesso')
linha = 1
coluna = 1
total_coluna = sheet.max_column - 1

for rows in sheet.iter_rows(min_row=3):

    for cell in rows:
        tmp = cell.value
        if tmp is None:
            tmp = "''"
        elif type(tmp) is str:
            tmp = "'" + tmp + "'"
        elif type(tmp) is int or type(tmp) is float:
            # tmp = str(int(sheet.cell(row=i+1, column=j + 1).value))
            tmp = "'" + str(int(tmp)) + "'"
        else:
            # tmp = str(sheet.cell(row=i+1, column=j + 1).value)
            tmp = "'" + tmp + "'"
        if coluna == total_coluna:
            data = data + tmp
            print(data)
            break
        else:
            data = data + tmp + ','
            coluna += 1


    coluna = 1
    database.insert_table_IES(data, linha)
    linha += 1
    data = ''

print(data)
